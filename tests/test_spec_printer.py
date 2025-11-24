"""Tests for the SpecPrinter class in dcmspec.spec_model."""
import gc
import logging
import pytest

from anytree import Node
from openpyxl import load_workbook

from dcmspec.spec_model import SpecModel
from dcmspec.spec_printer import SpecPrinter, SPECIAL_COLORS, LEVEL_COLORS

# Mock functions for testing
def mock_console_print_no_op(*args, **kwargs):
    """Mock console.print that does nothing.
    
    Args:
        *args: Positional arguments
        **kwargs: Keyword arguments 

    Returns:
        None

    """
    pass

def mock_table_add_row_capture_styles(styles_list):
    """Create a mock Table.add_row that captures styles.
    
    Args:
        styles_list: List to capture style values

    Returns:
        Mock function that captures styles

    """
    def _mock(*args, style=None, **kwargs):
        styles_list.append(style)
    return _mock

def mock_table_add_column_capture_widths(widths_list):
    """Create a mock Table.add_column that captures width arguments.
    
    Args:
        widths_list: List to capture width values
        
    Returns:
        Mock function that captures widths and calls original add_column

    """
    from rich.table import Table
    original_add_column = Table.add_column
    
    def _mock(self, header, width=None, **kwargs):
        widths_list.append(width)
        return original_add_column(self, header, width=width, **kwargs)
    return _mock

def mock_console_print_capture_text(outputs_list):
    """Create a mock console.print that captures text output.
    
    Args:
        outputs_list: List to capture text output
        
    Returns:
        Mock function that captures text output

    """
    def _mock(text, *args, **kwargs):
        outputs_list.append(text)
    return _mock

def mock_console_print_capture_styles(styles_list):
    """Create a mock console.print that captures style kwargs.
    
    Args:
        styles_list: List to capture style values
        
    Returns:
        Mock function that captures styles and calls original add_column

    """
    def _mock(text, *args, **kwargs):
        styles_list.append(kwargs.get('style'))
    return _mock

@pytest.fixture
def minimal_spec_model():
    """Create a minimal SpecModel with empty metadata and content nodes."""
    metadata = Node("metadata")
    metadata.header = []
    metadata.column_to_attr = {}
    content = Node("content")
    return SpecModel(metadata=metadata, content=content)

def add_standard_node(model, with_elem_tag=True):
    """Add a standard test node to the model's content tree and set typical metadata."""
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    node = Node("element1", parent=model.content)
    setattr(node, "elem_name", "Element1")
    if with_elem_tag:
        setattr(node, "elem_tag", "(0101,0010)")
    return node

def test_init_sets_model_and_logger(minimal_spec_model):
    """Test that SpecPrinter initializes model and logger correctly."""
    model = minimal_spec_model
    printer = SpecPrinter(model)
    assert printer.model is model
    assert isinstance(printer.logger, logging.Logger)

def test_init_raises_typeerror_for_bad_logger(minimal_spec_model):
    """Test that SpecPrinter raises TypeError if logger is not a Logger instance."""
    model = minimal_spec_model
    with pytest.raises(TypeError):
        SpecPrinter(model, logger="not_a_logger")

def test_print_tree_does_not_crash(monkeypatch, minimal_spec_model):
    """Test that print_tree can be called without error."""
    model = minimal_spec_model
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    printer.print_tree()

def test_print_tree_node_missing_attribute(monkeypatch, minimal_spec_model):
    """Test that print_tree handles nodes missing an attribute specified in attr_names."""
    model = minimal_spec_model
    add_standard_node(model, with_elem_tag=False)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    # Should not raise, should print empty string for missing attribute
    printer.print_tree(attr_names=["elem_name", "elem_tag"])

def test_print_tree_attr_names_string_and_list(monkeypatch, minimal_spec_model):
    """Test that print_tree works with attr_names as a string and as a list."""
    model = minimal_spec_model
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    # attr_names as string
    printer.print_tree(attr_names="elem_name")
    # attr_names as list
    printer.print_tree(attr_names=["elem_name", "elem_tag"])

def test_print_tree_attr_widths(monkeypatch, minimal_spec_model):
    """Test that print_tree applies attr_widths for padding/truncation."""
    model = minimal_spec_model
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    # attr_widths: elem_name padded to 5, elem_tag padded to 2
    printer.print_tree(attr_names=["elem_name", "elem_tag"], attr_widths=[5, 2])

def test_print_tree_no_color(monkeypatch, minimal_spec_model):
    """Test that print_tree does not apply color styles when colorize=False."""
    model = minimal_spec_model
    add_standard_node(model)
    printer = SpecPrinter(model)
    styles = []
    
    # Patch printer.console.print to capture the style attribute of the Text object for each print call
    monkeypatch.setattr(
        printer.console,
        "print",
        lambda text, *args, **kwargs: styles.append(getattr(text, "style", None)),
    )

    printer.print_tree(colorize=False)
    # All styles should be "default" or None (Rich may use None for no style)
    assert all(s in ("default", None, "") for s in styles), f"Unexpected styles: {styles}"
    
def test_print_table_does_not_crash(monkeypatch, minimal_spec_model):
    """Test that print_table can be called without error."""
    model = minimal_spec_model
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    printer.print_table()

def test_print_table_empty_header(monkeypatch, minimal_spec_model):
    """Test that print_table works when metadata.header is empty."""
    model = minimal_spec_model
    model.metadata.header = []  # Explicitly set header to empty for clarity
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    printer.print_table()

def test_print_table_node_missing_attribute(monkeypatch, minimal_spec_model):
    """Test that print_table handles nodes missing an attribute defined in column_to_attr."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    add_standard_node(model, with_elem_tag=False)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    # Should not raise, should print empty string for missing attribute
    printer.print_table()

def test_print_table_row_style(monkeypatch, minimal_spec_model):
    """Test that print_table sets the correct row style for include, title, and default nodes."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}

    # Include node
    include_node = Node("include_table_macro", parent=model.content)
    setattr(include_node, "elem_name", "Include Table Macro")

    # Title node
    title_node = Node("title_node", parent=model.content)
    setattr(title_node, "elem_name", "Module Title")
    # Patch _is_title to return True for title_node only
    model._is_title = lambda node: node is title_node

    # Standard node
    add_standard_node(model)

    printer = SpecPrinter(model)

    # Capture the styles passed to add_row
    styles = []

    # Patch Table.add_row
    monkeypatch.setattr("rich.table.Table.add_row", mock_table_add_row_capture_styles(styles))
    # Patch console.print to avoid output
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)

    printer.print_table(colorize=True)

    # The order is: include_node, title_node, default_node
    assert SPECIAL_COLORS["include"] in styles
    assert SPECIAL_COLORS["title"] in styles
    assert any(s in LEVEL_COLORS for s in styles)

def test_print_table_no_color(monkeypatch, minimal_spec_model):
    """Test that print_table sets style=None for all rows when colorize=False."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    # Add two standard nodes with different tags
    add_standard_node(model)
    node2 = Node("element2", parent=model.content)
    setattr(node2, "elem_name", "Element2")
    setattr(node2, "elem_tag", "(0101,0020)")
    printer = SpecPrinter(model)
    # Patch Table.add_row to track calls
    styles = []
    monkeypatch.setattr("rich.table.Table.add_row", mock_table_add_row_capture_styles(styles))
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    printer.print_table(colorize=False)
    # All add_row calls should have style=None
    assert all(s is None for s in styles)

def test_print_table_column_widths(monkeypatch, minimal_spec_model):
    """Test that print_table applies column_widths for column sizing."""
    model = minimal_spec_model
    node = add_standard_node(model)
    # Set headers AFTER add_standard_node to avoid them being overwritten
    model.metadata.header = ["Name", "Tag", "Type"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag", 2: "elem_type"}
    setattr(node, "elem_type", "1")
    
    printer = SpecPrinter(model)
    
    # Capture column widths
    column_widths_used = []
    from rich.table import Table
    monkeypatch.setattr(Table, "add_column", mock_table_add_column_capture_widths(column_widths_used))
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    
    printer.print_table(column_widths=[30, 15, 5])
    
    # Verify the widths were applied correctly
    assert column_widths_used == [30, 15, 5]

def test_print_table_column_widths_default(monkeypatch, minimal_spec_model):
    """Test that print_table uses default width 20 when column_widths is None."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    add_standard_node(model)
    
    printer = SpecPrinter(model)
    
    # Capture column widths
    column_widths_used = []
    from rich.table import Table
    monkeypatch.setattr(Table, "add_column", mock_table_add_column_capture_widths(column_widths_used))
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    
    printer.print_table()  # No column_widths specified
    
    # All columns should use default width of 20
    assert column_widths_used == [20, 20]

def test_print_table_column_widths_partial(monkeypatch, minimal_spec_model):
    """Test that print_table falls back to default width for unspecified columns."""
    model = minimal_spec_model
    node = add_standard_node(model)
    # Set headers AFTER add_standard_node to avoid them being overwritten
    model.metadata.header = ["Name", "Tag", "Type"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag", 2: "elem_type"}
    setattr(node, "elem_type", "1")
    
    printer = SpecPrinter(model)
    
    # Capture column widths
    column_widths_used = []
    from rich.table import Table
    monkeypatch.setattr(Table, "add_column", mock_table_add_column_capture_widths(column_widths_used))
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    
    # Only specify width for first two columns
    printer.print_table(column_widths=[25, 10])
    
    # First two columns use specified widths, third falls back to default
    assert column_widths_used == [25, 10, 20]

def test_print_csv_does_not_crash(monkeypatch, minimal_spec_model):
    """Test that print_csv can be called without error."""
    model = minimal_spec_model
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    printer.print_csv()

def test_print_csv_nominal(monkeypatch, minimal_spec_model):
    """Nominal CSV output: header + one data row with expected quoted values."""
    model = minimal_spec_model
    add_standard_node(model)  # sets header and column_to_attr
    printer = SpecPrinter(model)
    outputs = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_text(outputs))
    printer.print_csv()
    assert len(outputs) == 2, f"Unexpected output lines: {outputs}"
    assert outputs[0] == '"Name","Tag"'
    assert outputs[1] == '"Element1","(0101,0010)"'

def test_print_csv_multiline_field(monkeypatch, minimal_spec_model):
    """CSV preserves embedded newlines inside quoted fields."""
    model = minimal_spec_model
    node = add_standard_node(model)
    setattr(node, "elem_name", "Element1 line1\nline2")
    printer = SpecPrinter(model)
    outputs = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_text(outputs))
    printer.print_csv()
    assert outputs[0] == '"Name","Tag"'
    # Newline remains inside the quoted field
    assert outputs[1] == '"Element1 line1\nline2","(0101,0010)"'

def test_print_csv_empty_header(monkeypatch, minimal_spec_model):
    """Test that print_csv works when metadata.header is empty."""
    model = minimal_spec_model
    model.metadata.header = []
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    add_standard_node(model)
    printer = SpecPrinter(model)
    monkeypatch.setattr(printer.console, "print", mock_console_print_no_op)
    printer.print_csv()

def test_print_csv_empty_column_to_attr(monkeypatch, minimal_spec_model):
    """Test that print_csv handles empty column_to_attr correctly."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    add_standard_node(model)  # Adds a node and sets column_to_attr
    model.metadata.column_to_attr = {}  # Clear columns after adding the node
    printer = SpecPrinter(model)
    outputs = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_text(outputs))
    printer.print_csv()
    # Only header expected (rows have no columns -> empty row skipped)
    assert len(outputs) == 1

def test_print_csv_node_missing_attribute(monkeypatch, minimal_spec_model):
    """Test that print_csv handles nodes missing an attribute."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    add_standard_node(model, with_elem_tag=False)
    printer = SpecPrinter(model)
    outputs = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_text(outputs))
    printer.print_csv()
    # Header + one row
    assert len(outputs) == 2
    # Missing elem_tag should yield an empty quoted field
    assert outputs[1].endswith(',""') or ',""' in outputs[1]

def test_print_csv_row_style(monkeypatch, minimal_spec_model):
    """Test that print_csv sets the correct row style for different node types."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}

    # sourcery skip: extract-duplicate-method
    include_node = Node("include_table_macro", parent=model.content)
    setattr(include_node, "elem_name", "Include Item")
    setattr(include_node, "elem_tag", "(0101,1001)")

    title_node = Node("title_node", parent=model.content)
    setattr(title_node, "elem_name", "Title Item")
    setattr(title_node, "elem_tag", "(0101,1002)")
    model._is_title = lambda n: n is title_node  # Patch title logic

    add_standard_node(model)

    printer = SpecPrinter(model)
    styles = []

    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_styles(styles))

    printer.print_csv(colorize=True)

    row_styles = [s for s in styles[1:] if s is not None]
    assert SPECIAL_COLORS["include"] in row_styles
    assert SPECIAL_COLORS["title"] in row_styles
    assert any(s in LEVEL_COLORS for s in row_styles)

def test_print_csv_no_color(monkeypatch, minimal_spec_model):
    """Test that print_csv sets style=None for all rows when colorize=False."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    add_standard_node(model)
    printer = SpecPrinter(model)
    styles = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_styles(styles))
    printer.print_csv(colorize=False)
    # Header + data rows all should have style None
    assert all(s is None for s in styles), f"Unexpected non-None styles: {styles}"

def test_print_csv_quote_escaping(monkeypatch, minimal_spec_model):
    """Test that print_csv properly escapes quotes in CSV output."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    node = add_standard_node(model)
    setattr(node, "elem_name", 'Element "Quoted" Name')
    printer = SpecPrinter(model)
    outputs = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_text(outputs))
    printer.print_csv()
    # Row should have doubled internal quotes
    data_rows = outputs[1:]
    assert any('Element ""Quoted"" Name' in r for r in data_rows)

def test_print_csv_skips_empty_rows(monkeypatch, minimal_spec_model):
    """Test that print_csv skips rows with only whitespace."""
    model = minimal_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    node = add_standard_node(model)
    setattr(node, "elem_name", " ")
    setattr(node, "elem_tag", " ")
    printer = SpecPrinter(model)
    outputs = []
    monkeypatch.setattr(printer.console, "print", mock_console_print_capture_text(outputs))
    printer.print_csv()
    # Only header printed because row is empty (whitespace)
    assert len(outputs) == 1

def test_init_with_output_file(minimal_spec_model, tmp_path):
    """Test that SpecPrinter initializes with an output file path."""
    output_file = tmp_path / "output.txt"
    printer = SpecPrinter(minimal_spec_model, output=str(output_file))
    assert printer.output == str(output_file)
    assert printer.console.file is not None
    # Clean up
    del printer

def test_print_tree_to_file(minimal_spec_model, tmp_path):
    """Test that print_tree writes plain text to file without colors."""
    output_file = tmp_path / "tree_output.txt"
    add_standard_node(minimal_spec_model)
    printer = SpecPrinter(minimal_spec_model, output=str(output_file))
    printer.print_tree(attr_names="elem_name", colorize=True)
    del printer  # Ensure file is closed
    
    content = output_file.read_text(encoding="utf-8")
    assert "Element1" in content
    # Tree drawing characters should be present
    assert "└──" in content or "├──" in content
    # Verify no ANSI color codes (RGB sequences won't be in plain text)
    assert "\x1b[" not in content

def test_print_table_to_file(minimal_spec_model, tmp_path):
    """Test that print_table writes plain text to file without colors."""
    output_file = tmp_path / "table_output.txt"
    add_standard_node(minimal_spec_model)
    printer = SpecPrinter(minimal_spec_model, output=str(output_file))
    printer.print_table(colorize=True)
    del printer
    
    content = output_file.read_text()
    assert "Element1" in content
    assert "(0101,0010)" in content
    assert "\x1b[" not in content

def test_print_csv_to_file(minimal_spec_model, tmp_path):
    """Test that print_csv writes CSV to file without colors."""
    output_file = tmp_path / "csv_output.csv"
    add_standard_node(minimal_spec_model)
    printer = SpecPrinter(minimal_spec_model, output=str(output_file))
    printer.print_csv(colorize=True)
    del printer
    
    content = output_file.read_text()
    lines = content.strip().split("\n")
    assert len(lines) == 2  # Header + 1 data row
    assert lines[0] == '"Name","Tag"'
    assert lines[1] == '"Element1","(0101,0010)"'
    assert "\x1b[" not in content

def test_output_file_closed_on_delete(minimal_spec_model, tmp_path):
    """Test that output file is properly closed when printer is deleted."""
    output_file = tmp_path / "closed_test.txt"
    printer = SpecPrinter(minimal_spec_model, output=str(output_file))
    file_obj = printer.console.file
    del printer
    gc.collect()  # Force garbage collection so __del__ runs
    assert file_obj.closed

def test_print_xlsx_basic(minimal_spec_model, tmp_path):
    """Basic XLSX export: header and one data row written."""
    add_standard_node(minimal_spec_model)
    output_file = tmp_path / "out.xlsx"
    printer = SpecPrinter(minimal_spec_model)
    printer.print_xlsx(output=str(output_file), colorize=False)

    wb = load_workbook(str(output_file))
    ws = wb.active
    # Header values
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=1, column=2).value == "Tag"
    # Data row values
    assert ws.cell(row=2, column=1).value == "Element1"
    assert ws.cell(row=2, column=2).value == "(0101,0010)"
    # No fill when colorize=False
    assert ws.cell(row=2, column=1).fill.patternType in (None, "solid")  # may default to solid without color

def test_print_xlsx_colorize_and_column_widths(minimal_spec_model, tmp_path):
    """XLSX export applies color fill and column widths."""
    node = add_standard_node(minimal_spec_model)
    # Add a third column
    minimal_spec_model.metadata.header = ["Name", "Tag", "Type"]
    minimal_spec_model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag", 2: "elem_type"}
    setattr(node, "elem_type", "1")
    output_file = tmp_path / "styled.xlsx"
    printer = SpecPrinter(minimal_spec_model)
    printer.print_xlsx(output=str(output_file), colorize=True, column_widths=[25, 15, 8])

    wb = load_workbook(str(output_file))
    ws = wb.active

    # Verify column widths (rounded by Excel; allow tolerance)
    assert abs(ws.column_dimensions['A'].width - 25) < 0.5
    assert abs(ws.column_dimensions['B'].width - 15) < 0.5
    assert abs(ws.column_dimensions['C'].width - 8) < 0.5

    # Color fill applied: depth 1 → LEVEL_COLORS[0] rgb(176,224,230) → hex B0E0E6 (openpyxl stores ARGB)
    fill_rgb = ws.cell(row=2, column=1).fill.start_color.rgb
    assert fill_rgb is not None
    assert fill_rgb.endswith("B0E0E6")





