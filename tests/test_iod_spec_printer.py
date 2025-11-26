"""Tests for the IODSpecPrinter class in dcmspec.iod_spec_printer."""
import pytest

from anytree import Node
from openpyxl import load_workbook

from dcmspec.iod_spec_printer import IODSpecPrinter
from dcmspec.spec_model import SpecModel

@pytest.fixture
def iod_spec_model():
    """Fixture for a minimal IOD SpecModel with a module and attribute node."""
    metadata = Node("metadata")
    metadata.header = ["Attr1", "Attr2"]
    metadata.column_to_attr = {0: "attr1", 1: "attr2"}
    content = Node("content")
    # Add a module node
    module_node = Node("module1", parent=content)
    setattr(module_node, "module", "Patient")
    setattr(module_node, "usage", "M")
    # Add an attribute node under the module node
    attr_node = Node("attr", parent=module_node)
    setattr(attr_node, "attr1", "Value1")
    setattr(attr_node, "attr2", "Value2")
    # Add helpers for color logic
    model = SpecModel(metadata=metadata, content=content)
    model._is_include = lambda node: False
    model._is_title = lambda node: False
    return model

def set_attr1_with_quote(model):
    """Set attr1 to a value with a quote for all attribute nodes in the model."""
    for node in model.content.descendants:
        if hasattr(node, "attr1"):
            node.attr1 = 'Val"ue1'

def _capture_print(outputs):
    """Return a function that appends the first argument (text) to outputs when called.

    Useful for monkeypatching printer.console.print to capture printed text in tests.
    """
    def capture_print(text, *args, **kwargs):
        outputs.append(text)
    return capture_print

def _capture_style(styles, orig_print):
    """Return a function that appends the 'style' kwarg to styles and calls the original print.

    Useful for monkeypatching printer.console.print to capture style arguments in tests.
    """
    def capture_style(*args, **kwargs):
        style = kwargs.get("style", None)
        styles.append(style)
        return orig_print(*args, **kwargs)
    return capture_style

def test_print_table_prints_module_title_and_attr(iod_spec_model):
    """Test that module title and attribute values are present in the output."""
    printer = IODSpecPrinter(iod_spec_model)
    # Using printer.console.capture() is sufficient here because we only need to capture the rendered text output.
    # Patching is not necessary unless we want to inspect styles or other print arguments.
    with printer.console.capture() as capture:
        printer.print_table(colorize=False)
    result = capture.get()
    # Check for module title and attribute values
    assert "Patient Module (M)" in result
    assert "Value1" in result
    assert "Value2" in result

def test_print_table_column_widths(iod_spec_model):
    """Test that print_table applies column_widths for column sizing."""
    printer = IODSpecPrinter(iod_spec_model)
    # Capture the console output
    with printer.console.capture() as capture:
        printer.print_table(column_widths=[5, 2], colorize=False)
    result = capture.get()
    # Check that the headers and values are truncated/padded according to specified widths.
    # Rich will pad or truncate headers and cell values to fit the specified column widths.
    # If the content is too wide for the column, Rich truncates it and appends a Unicode ellipsis (…).
    # If the content is shorter than the column width, Rich pads it with spaces.
    assert "Attr1" in result
    assert "A…" in result  # "Attr2" truncated to 2 chars
    assert "Valu…" in result  # "Value1" truncated to 5 chars
    assert "Va" in result  # "Value2" truncated to 2 chars

def test_print_table_no_color_styles(iod_spec_model, monkeypatch):
    """Test that print_table does not apply color styles when colorize=False."""
    printer = IODSpecPrinter(iod_spec_model)
    styles = []
    # Patch printer.console.print to intercept and record the 'style' argument for each print call.
    monkeypatch.setattr(printer.console, "print", _capture_style(styles, printer.console.print))
    printer.print_table(colorize=False)
    # All styles should be None or default
    assert all(s in (None, "default", "") for s in styles)

def test_print_table_no_color_when_output_set(monkeypatch, iod_spec_model, tmp_path):
    """Test that print_table sets style=None for all rows when output is set (writing to file)."""
    # Prepare model
    model = iod_spec_model
    model.metadata.header = ["Name", "Tag"]
    model.metadata.column_to_attr = {0: "elem_name", 1: "elem_tag"}
    
    # Add two nodes
    node2 = Node("element2", parent=model.content)
    setattr(node2, "elem_name", "Element2")
    setattr(node2, "elem_tag", "(0101,0020)")
    
    # Simulate output being set
    output_file = tmp_path / "output.txt"
    printer = IODSpecPrinter(model, output=str(output_file))
    
    # Capture styles passed to Table.add_row
    styles = []
    monkeypatch.setattr(printer.console, "print", _capture_style(styles, printer.console.print))
    
    # Call method with colorize=True (should be ignored because output is set)
    printer.print_table(colorize=True)
    
    # Assert that all styles are None (no color applied)
    assert all(s is None for s in styles), f"Expected no color styles, got {styles}"

def test_print_tree_prints_module_title_and_attr(iod_spec_model):
    """Test that print_tree outputs the module title and attribute values."""
    printer = IODSpecPrinter(iod_spec_model)
    # Capture the console output
    with printer.console.capture() as capture:
        printer.print_tree(attr_names=["attr1", "attr2"], colorize=False)
    result = capture.get()
    # Check for module title and attribute values
    assert "Patient Module (M)" in result
    assert "Value1" in result
    assert "Value2" in result

def test_print_tree_handles_attr_names_and_widths(iod_spec_model):
    """Test that print_tree respects attr_names and attr_widths."""
    printer = IODSpecPrinter(iod_spec_model)
    # Capture the console output
    with printer.console.capture() as capture:
        printer.print_tree(attr_names=["attr1", "attr2"], attr_widths=[2, 3], colorize=False)
    result = capture.get()
    # Check for module title and attribute values
    # Should be truncated to, respectively, 2 and 3 characters
    assert "Va Val" in result

def test_print_tree_no_color_styles(iod_spec_model, monkeypatch):
    """Test that print_tree does not apply color styles when colorize=False."""
    printer = IODSpecPrinter(iod_spec_model)
    styles = []
    # Patch printer.console.print to intercept and record the 'style' argument for each print call.
    monkeypatch.setattr(printer.console, "print", _capture_style(styles, printer.console.print))
    printer.print_tree(colorize=False)
    assert all(s in ("default", None, "") for s in styles)

def test_print_tree_no_color_when_output_set(monkeypatch, iod_spec_model, tmp_path):
    """Test that print_tree does not apply color styles when output is set (writing to file)."""
    model = iod_spec_model
    
    # Simulate output being set (e.g., writing to a file)
    output_file = tmp_path / "output.txt"
    printer = IODSpecPrinter(model, output=str(output_file))
    
    styles = []
    
    # Patch printer.console.print to capture the style attribute
    monkeypatch.setattr(
        printer.console,
        "print",
        lambda text, *args, **kwargs: styles.append(getattr(text, "style", None)),
    )
    
    printer.print_tree(colorize=True)  # Even if colorize=True, output disables it
    
    # Assert that no style was applied
    assert all(style in (None, '') for style in styles), f"Expected no color styles, got {styles}"


def test_print_csv_prints_module_title_and_attr(iod_spec_model, monkeypatch):
    """Test that print_csv outputs the module title and attribute values as CSV."""
    printer = IODSpecPrinter(iod_spec_model)
    outputs = []
    # Patch printer.console.print to capture printed lines
    monkeypatch.setattr(printer.console, "print", _capture_print(outputs))
    printer.print_csv(colorize=False)
    # The first line is the header, the second is the module title row, the third is the attribute row
    assert outputs[0] == '"Attr1","Attr2"'
    assert '"Patient Module (M)",""' in outputs  # Module title row
    assert '"Value1","Value2"' in outputs        # Attribute row

def test_print_csv_escapes_quotes(iod_spec_model, monkeypatch):
    """Test that print_csv escapes quotes in CSV output."""
    set_attr1_with_quote(iod_spec_model)
    printer = IODSpecPrinter(iod_spec_model)
    outputs = []
    # Patch printer.console.print to capture printed lines
    monkeypatch.setattr(printer.console, "print", _capture_print(outputs))
    printer.print_csv(colorize=False)
    assert any('Val""ue1' in line for line in outputs)

def test_print_csv_no_color_styles(iod_spec_model, monkeypatch):
    """Test that print_csv does not apply color styles when colorize=False."""
    printer = IODSpecPrinter(iod_spec_model)
    styles = []
    # Patch printer.console.print to capture printed lines
    monkeypatch.setattr(printer.console, "print", _capture_style(styles, printer.console.print))
    printer.print_csv(colorize=False)
    # All styles should be None
    assert all(s is None for s in styles)

def test_print_csv_no_color_when_output_set(monkeypatch, iod_spec_model, tmp_path):
    """Test that print_csv sets style=None for all rows when output is set (writing to file)."""
    # Prepare model
    model = iod_spec_model

    # Simulate output being set
    output_file = tmp_path / "output.csv"
    printer = IODSpecPrinter(model, output=str(output_file))

    # Capture styles from console.print
    styles = []
    monkeypatch.setattr(printer.console, "print", _capture_style(styles, printer.console.print))

    # Call method with colorize=True (should be ignored because output is set)
    printer.print_csv(colorize=True)

    # Header + data rows should all have style=None
    assert all(s in (None, '') for s in styles), f"Unexpected non-None styles: {styles}"

def test_print_xlsx_basic(iod_spec_model, tmp_path):
    """Test that print_xlsx writes header and one data row to a worksheet named after the module."""
    output_file = tmp_path / "iod_out.xlsx"
    printer = IODSpecPrinter(iod_spec_model, output=str(output_file))
    printer.print_xlsx(colorize=False)

    wb = load_workbook(str(output_file))
    # Should be a single worksheet named "Patient"
    assert "Patient" in wb.sheetnames
    ws = wb["Patient"]
    # Header values
    assert ws.cell(row=1, column=1).value == "Attr1"
    assert ws.cell(row=1, column=2).value == "Attr2"
    # Data row values
    assert ws.cell(row=2, column=1).value == "Value1"
    assert ws.cell(row=2, column=2).value == "Value2"
    # No fill when colorize=False
    assert ws.cell(row=2, column=1).fill.patternType in (None, "solid")

def test_print_xlsx_column_widths(iod_spec_model, tmp_path):
    """Test that print_xlsx applies column widths to the worksheet."""
    output_file = tmp_path / "iod_colwidths.xlsx"
    printer = IODSpecPrinter(iod_spec_model, output=str(output_file))
    printer.print_xlsx(colorize=False, column_widths=[15, 8])

    wb = load_workbook(str(output_file))
    ws = wb["Patient"]
    # openpyxl may round column widths, so allow a small tolerance
    assert abs(ws.column_dimensions['A'].width - 15) < 0.5
    assert abs(ws.column_dimensions['B'].width - 8) < 0.5

def test_print_xlsx_colorize(iod_spec_model, tmp_path):
    """Test that print_xlsx applies color fill to attribute rows when colorize=True."""
    output_file = tmp_path / "iod_color.xlsx"
    printer = IODSpecPrinter(iod_spec_model, output=str(output_file))
    printer.print_xlsx(colorize=True)

    wb = load_workbook(str(output_file))
    ws = wb["Patient"]
    # Data row should have a fill color (LEVEL_COLORS[0])
    fill_rgb = ws.cell(row=2, column=1).fill.start_color.rgb
    assert fill_rgb is not None
    # LEVEL_COLORS[0] is "rgb(176,224,230)" → hex "B0E0E6"
    assert fill_rgb == "FF87CEFA"

def test_print_xlsx_multiple_modules(iod_spec_model, tmp_path):
    """Test that print_xlsx creates a worksheet for each module, including duplicates."""
    content = iod_spec_model.content

    # Add a second module with a different name
    module_node2 = Node("module2", parent=content)
    setattr(module_node2, "module", "Series")
    setattr(module_node2, "usage", "M")
    attr_node2 = Node("attr2", parent=module_node2)
    setattr(attr_node2, "attr1", "Series1")
    setattr(attr_node2, "attr2", "Series2")

    # Add a third module with the same name as the first
    module_node3 = Node("module3", parent=content)
    setattr(module_node3, "module", "Patient")
    setattr(module_node3, "usage", "M")
    attr_node3 = Node("attr3", parent=module_node3)
    setattr(attr_node3, "attr1", "PatientX")
    setattr(attr_node3, "attr2", "PatientY")

    output_file = tmp_path / "iod_multi_modules.xlsx"
    printer = IODSpecPrinter(iod_spec_model, output=str(output_file))
    printer.print_xlsx(colorize=False)

    wb = load_workbook(str(output_file))
    # Should have three sheets: "Patient", "Series", "Patient_1"
    assert "Patient" in wb.sheetnames
    assert "Series" in wb.sheetnames
    assert "Patient_1" in wb.sheetnames

    ws_patient = wb["Patient"]
    ws_series = wb["Series"]
    ws_patient_1 = wb["Patient_1"]

    # Check that the data is in the correct sheets
    assert ws_patient.cell(row=2, column=1).value == "Value1"
    assert ws_series.cell(row=2, column=1).value == "Series1"
    assert ws_patient_1.cell(row=2, column=1).value == "PatientX"

