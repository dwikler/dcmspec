"""Printer class for specification model in dcmspec.

Provides the SpecPrinter class for printing DICOM specification models (SpecModel)
to a file or standard output.
"""

from typing import Generator, Optional, List, Tuple, Union
import logging

from rich.console import Console
from rich.table import Table, box
from rich.text import Text
from anytree import Node, RenderTree, PreOrderIter
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

LEVEL_COLORS = [
    "rgb(176,224,230)",   # Root: Powder Blue
    "rgb(135,206,250)",   # Level 1: Sky Blue
    "rgb(0,191,255)",     # Level 2: Deep Sky Blue
    "rgb(30,144,255)",    # Level 3: Dodger Blue
    "rgb(65,105,225)",    # Level 4: Royal Blue
    "rgb(106,90,205)",    # Level 5: Slate Blue
    "rgb(123,104,238)",   # Level 7: Medium Slate Blue
    "rgb(147,112,219)",   # Level 8: Medium Purple
]

SPECIAL_COLORS = {
    "include": "rgb(255,255,135)",  # Yellow for include nodes
    "title": "rgb(255,0,255)",    # Magenta for title nodes
}

class SpecPrinter:
    """Printer for DICOM specification models.

    Provides methods to render a SpecModel in multiple formats:
    - Hierarchical tree (ASCII)
    - Flat table (ASCII, using Rich for styling)
    - CSV (plain text)

    Output can be directed to the console (default) or to a file by specifying
    an output path when initializing the printer. Rich formatting is used for
    tree and table views when writing to the console; when writing to a file,
    plain text is used.

    Responsibilities:
    - Encapsulates all presentation logic for SpecModel.
    - Supports colorized output for better readability.
    - Handles output destination internally (stdout or file).

    Args:
        model (SpecModel): The specification model to print.
        output (Optional[str]): Path to an output file. If None, prints to stdout.
        logger (Optional[logging.Logger]): Logger instance for debug/info messages.

    """

    def __init__(self, model: object, output: Optional[str] = None, logger: Optional[logging.Logger] = None) -> None:
        """Initialize the printer with a specification model and optional output destination.

        Args:
            model (object): An instance of SpecModel to render.
            output (Optional[str]): Path to an output file. If None, defaults to stdout.
            logger (Optional[logging.Logger]): Logger instance for debug/info messages.
            
        """
        self.model = model
        self.output = output
        self.console = None  # ensure attribute exists even if validation raises

        if logger is not None and not isinstance(logger, logging.Logger):
            raise TypeError("logger must be an instance of logging.Logger or None")
        self.logger = logger or logging.getLogger(self.__class__.__name__)

        self.console = Console(
            highlight=False,
            file=open(output, "w", encoding="utf-8") if output else None,
            force_terminal=not output,
            no_color=bool(output)
        )

    def print_tree(
        self,
        attr_names: Optional[Union[str, List[str]]] = None,
        attr_widths: Optional[List[int]] = None,
        colorize: bool = False,
    ) -> None:
        """Print the specification model as a hierarchical tree to the console or file.

        Args:
            attr_names (Optional[Union[str, list[str]]]): Attribute name(s) to display for each node.
                If None, only the node's name is displayed.
                If a string, displays that single attribute.
                If a list of strings, displays all specified attributes.
            attr_widths (Optional[list[int]]): List of widths for each attribute in attr_names.
                If provided, each attribute will be padded/truncated to the specified width.
            colorize (bool): Whether to colorize the output by node depth. Ignored when writing to file.

        Returns:
            None

        Example:
            ```python
            # This will nicely align the tag, type, and name values in the tree output:
            printer.print_tree(attr_names=["elem_tag", "elem_type", "elem_name"], attr_widths=[11, 2, 64])
            ```

        """
        # Disable colorization if writing to file
        if self.output:
            colorize = False

        for pre, fill, node in RenderTree(self.model.content):
            pre_text = Text(pre)
            attr_text = self._format_tree_row(node, attr_names, attr_widths)
            row_style = self._get_tree_row_style(node, colorize)
            node_text = Text(attr_text, style=row_style)
            self.console.print(pre_text + node_text)

        if self.console.file:
            self.console.file.flush()  # Ensure data is written immediately

    def print_table(self, column_widths: Optional[List[int]] = None, colorize: bool = False) -> None:
        """Print the specification model as an ascii table to the console or file.

        Traverses the content tree and prints each node's attributes in a flat table,
        using column headers from the metadata node. Optionally colorizes rows.


        Args:
            column_widths (Optional[List[int]]): List of widths for each column's **content**.
                These widths do not include borders or padding added by Rich.
                If provided, each column will be set to the specified content width.
                If None, all columns default to width 20.            
                If the list is shorter than the number of columns, remaining columns default to width 20.

            colorize (bool): Whether to colorize the output by node depth. Ignored when writing to file.

        Returns:
            None
            
        """
        # Disable colorization if writing to file
        if self.output:
            colorize = False
            
        table = Table(show_header=True, header_style="bold magenta", show_lines=True, box=box.ASCII_DOUBLE_HEAD)

        # Define the columns using the extracted headers
        for i, header in enumerate(self.model.metadata.header):
            width = column_widths[i] if column_widths and i < len(column_widths) else 20
            table.add_column(header, width=width)

        # Add rows to the table
        for row, row_style in self._iterate_rows(colorize):
            table.add_row(*row, style=row_style)

        self.console.print(table)
        
        if self.console.file:
            self.console.file.flush()  # Ensure data is written immediately

    def print_csv(self, colorize: bool = False) -> None:
        """Print the specification model as CSV to the console or file.

        Traverses the content tree and prints each node's attributes in CSV format,
        with column headers from the metadata node. Optionally colorizes rows.

        Args:
            colorize (bool): Whether to colorize the output by node depth. Ignored when writing to file.

        Returns:
            None
            
        """
        # Disable colorization if writing to file
        if self.output:
            colorize = False
            
        # Print CSV header
        header_row = ",".join(f'"{h}"' for h in self.model.metadata.header)
        self.console.print(header_row)

        # Add data rows
        for row, row_style in self._iterate_rows(colorize):
            # Escape quotes inside each field by doubling them 
            # (e.g., Include Table 10.29-1 "UDI Macro Attributes"→ Include Table 10.29-1 ""UDI Macro Attributes""),
            # then wrap the entire field in quotes for proper CSV formatting
            csv_row = ",".join(f'"{cell.replace(chr(34), chr(34) + chr(34))}"' for cell in row)
            self.console.print(csv_row, style=row_style)

        if self.console.file:
            self.console.file.flush()  # Ensure data is written immediately

    def print_xlsx(self, column_widths: Optional[List[int]] = None, colorize: bool = False) -> None:
        """Print the specification model to an OOXML format Excel (.xlsx) file.

        Traverses the content tree and writes each node's attributes into an Excel sheet,
        with column headers from the metadata node. Handles newlines and applies background
        colorization using the same color scheme as console output (LEVEL_COLORS).

        Args:
            column_widths (list): Optional list of column widths for Excel columns.
            colorize (bool): Whether to apply color styling to cell backgrounds.

        Returns:
            None

            
        Note:
            All values are written as text to preserve DICOM identifiers and prevent unintended
            numeric conversion. Excel may still display a 'Number stored as text' warning for
            values that look numeric. This is expected and harmless. To suppress it:
            - On Windows: File → Options → Formulas → Error Checking → Uncheck "Numbers stored as text".
            - On macOS: Excel → Preferences → Error Checking → Uncheck "Numbers formatted as text".

        """
        if not self.output:
            raise ValueError("Output file path must be specified when constructing SpecPrinter for print_xlsx).")

        header_style, data_style = self._create_styles()
        wb = self._setup_workbook()
        ws = self._setup_worksheet(wb, "Specification")    

        self._write_headers(ws, self.model.metadata.header, header_style)
        if column_widths:
            self._set_column_widths(ws, column_widths)

        self._write_data_rows(ws, self._iterate_rows(colorize=colorize), data_style, colorize)
        wb.save(self.output)

    def _create_styles(self) -> Tuple[dict, dict]:
        border = Border(
            left=Side(style="thin", color="B3B3B3"),
            right=Side(style="thin", color="B3B3B3"),
            top=Side(style="thin", color="B3B3B3"),
            bottom=Side(style="thin", color="B3B3B3")
        )
        alignment = Alignment(wrap_text=True, vertical="top")
        header_style = {
            "font": Font(bold=True, name="Consolas"),
            "alignment": alignment,
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "border": border
        }
        data_style = {
            "font": Font(name="Consolas"),
            "alignment": alignment,
            "border": border
        }
        return header_style, data_style

    def _setup_workbook(self) -> Workbook:
        wb = Workbook()
        # Remove the default sheet if it exists
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        return wb
    
    def _setup_worksheet(self, wb: Workbook, title: str) -> Worksheet:
        return wb.create_sheet(title=title)

    def _write_headers(self, ws: Worksheet, headers: List[str], style: dict) -> None:
        ws.append(headers)
        for cell in ws[1]:
            self._apply_style(cell, style)

    def _set_column_widths(self, ws: Worksheet, widths: List[int]) -> None:
        from openpyxl.utils import get_column_letter
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _write_data_rows(
            self, 
            ws: Worksheet, 
            rows: Generator[Tuple[List[str], Optional[str]], None, None], 
            style: dict, 
            colorize: bool
            ) -> None:
        for row, row_style in rows:
            current_row = ws.max_row + 1
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.number_format = '@'
                cell.value = str(value)
                self._apply_style(cell, style)
                if colorize and row_style:
                    cell.fill = PatternFill(start_color=self._rgb_to_hex(row_style),
                                            end_color=self._rgb_to_hex(row_style),
                                            fill_type="solid")
                    
    @staticmethod
    def _apply_style(cell, style_dict) -> None:
        """Apply a style dictionary (font, alignment, fill, border) to a cell."""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)

    @staticmethod
    def _rgb_to_hex(rgb_str: str) -> str:
        """Convert an RGB string like 'rgb(255,255,0)' to ARGB hex format (e.g., 'FFFF00FF')."""
        rgb = rgb_str.replace("rgb(", "").replace(")", "").split(",")
        # Prepend 'FF' for opaque alpha channel as openpyxl expects ARGB
        return "FF{:02X}{:02X}{:02X}".format(*map(int, rgb))

    def _iterate_rows(self, colorize: bool = False) -> Generator[Tuple[List[str], Optional[str]], None, None]:
        """Generate rows from the model tree with optional styling.

        Args:
            colorize (bool): Whether to apply color styling to rows.

        Yields:
            tuple: (row_data, row_style) where row_data is a list of cell values
                    and row_style is an RGB color string (e.g., "rgb(255,255,0)") 
                    from LEVEL_COLORS or SPECIAL_COLORS, or None if not colorized.

        """
        for node in PreOrderIter(self.model.content):
            # skip the root node
            if node.name == "content":
                continue
            
            row = [str(getattr(node, attr, "") or "") for attr in self.model.metadata.column_to_attr.values()]
            # Skip row if all values are empty or whitespace
            if all(not str(cell).strip() for cell in row):
                continue
            row_style = None
            if colorize:
                if self.model._is_include(node):
                    row_style = SPECIAL_COLORS["include"]
                elif self.model._is_title(node):
                    row_style = SPECIAL_COLORS["title"]
                else:
                    row_style = LEVEL_COLORS[(node.depth - 1) % len(LEVEL_COLORS)]
            
            yield row, row_style

    def _format_tree_row(
            self, 
            node: Node, 
            attr_names: Optional[Union[str, List[str]]] = None, 
            attr_widths: Optional[List[int]] = None
            ) -> str:
        """Format the text for a node in tree output.

        Args:
            node: The node to display.
            attr_names (Optional[Union[str, list[str]]]): Attribute name(s) to display for each node.
            attr_widths (Optional[list[int]]): List of widths for each attribute in attr_names.

        Returns:
            str: The formatted text for the node.

        """
        if attr_names is None:
            # Just show the node name, safely handle missing name attribute
            return str(node.name)
        # Ensure attr_names is a list
        attr_names_list = [attr_names] if isinstance(attr_names, str) else attr_names
        # Collect attribute values, replacing None with empty string
        raw_values = []
        for attr in attr_names_list:
            value = getattr(node, attr, None)
            raw_values.append("" if value is None else str(value))

        if attr_widths:
            # Apply padding/truncation
            padded_values = []
            for v, w in zip(raw_values, attr_widths):
                if w is None:
                    padded_values.append(v)
                else:
                    padded_values.append(v.ljust(w)[:w])
            values = padded_values
        else:
            values = raw_values

        # Join values into a single string and remove leading spaces
        return " ".join(values).lstrip()

    def _get_tree_row_style(self, node: Node, colorize: bool = False) -> Optional[str]:
        """Determine the style for a node in tree output.

        Args:
            node: The node to display.
            colorize (bool): Whether to colorize the output by node depth.

        Returns:
            str or None: The style string for the node, or None if not colorized.
            
        """
        if not colorize:
            return "default"
        if hasattr(self.model, "_is_include") and self.model._is_include(node):
            return SPECIAL_COLORS["include"]
        elif hasattr(self.model, "_is_title") and self.model._is_title(node):
            return SPECIAL_COLORS["title"]
        else:
            return LEVEL_COLORS[(node.depth - 1) % len(LEVEL_COLORS)]

    def __del__(self) -> None:
        """Close output file when the printer is deleted, if opened."""
        output = getattr(self, "output", None)
        console = getattr(self, "console", None)
        file_obj = getattr(console, "file", None) if console else None
        if output and file_obj:
            try:
                if not file_obj.closed:
                    file_obj.close()
            except Exception:
                pass
